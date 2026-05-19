"""
ISX OHLCV Backfill
==================
Crawls all pages of the ISX portal market reports, downloads every daily
trading XLSX, parses OHLCV per company, and writes data/ohlcv.json.

Format: {"YYYY-MM-DD": {"SYM": {"o":f, "h":f, "l":f, "c":f, "v":i, "d":i}, ...}, ...}

Also appended to data/hist.json (the 's' short-history dict used by the site).
"""
import requests, json, os, re, datetime, time
import openpyxl
from io import BytesIO

BASE = 'http://www.isx-iq.net/isxportal'
DOMAIN = 'http://www.isx-iq.net'
H = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}

# ── Step 1: Crawl all pages and collect (date, xlsx_url) pairs ────────────────

def get_all_daily_xlsx():
    sess = requests.Session()
    sess.headers.update(H)
    results = []   # list of (date_iso, url)
    seen_urls = set()

    # Page 1
    r = sess.get(f'{BASE}/portal/uploadedFilesList.html', timeout=20)
    parse_page(r.text, results, seen_urls)

    # Find total pages from "last page" link (p=N where N is max)
    all_page_nums = re.findall(r'd-447146-p=(\d+)', r.text)
    total_pages = max((int(n) for n in all_page_nums), default=1)
    print(f'  Total pages: {total_pages}')

    # Crawl every page using the session (preserves jsessionid)
    for pnum in range(2, total_pages + 1):
        # Reuse the jsessionid-embedded URL pattern from page 1
        sid_m = re.search(r'uploadedFilesList\.html;(jsessionid=[A-Z0-9]+)\?', r.text)
        sid = sid_m.group(1) + '?' if sid_m else '?'
        url = f'{BASE}/portal/uploadedFilesList.html;{sid}d-447146-p={pnum}'
        print(f'  Scraping page {pnum}/{total_pages}...')
        r = sess.get(url, timeout=20)
        parse_page(r.text, results, seen_urls)
        time.sleep(0.3)

    return results

def parse_page(html, results, seen_urls):
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
    for row in rows:
        if 'files/' not in row or '.xlsx' not in row:
            continue
        link_m = re.search(r'href="(/isxportal/files/([^"]+\.xlsx))"', row)
        date_m = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', row)
        if not link_m or not date_m:
            continue
        path, fname = link_m.group(1), link_m.group(2)
        # Skip non-daily: Arabic-prefixed = weekly/news, starts with digit = daily
        # Arabic Unicode range starts at ؀
        if fname and ord(fname[0]) > 0x05FF:
            continue
        # Skip obvious non-daily patterns (monthly report filenames have 'monthly' etc.)
        fname_lower = fname.lower()
        if any(k in fname_lower for k in ['monthly', 'quarter', 'yearly', 'annual']):
            continue
        url = DOMAIN + path
        if url in seen_urls:
            continue
        seen_urls.add(url)
        # Parse date DD/MM/YYYY → YYYY-MM-DD
        d, m, y = date_m.group(1).split('/')
        date_iso = f'{y}-{m}-{d}'
        results.append((date_iso, url))
    return results

# ── Step 2: Download + parse each XLSX ───────────────────────────────────────

def parse_xlsx(content, date_iso):
    """
    Returns dict: {sym: {o, h, l, c, v, d}} for all traded companies.
    Sheet 'Bullient ' columns (0-indexed):
      0=blank, 1=Company Name, 2=Code, 3=Open, 4=High, 5=Low,
      6=AvgPrice, 7=PrevAvg, 8=Close, 9=PrevClose, 10=Chg%, 11=Deals, 12=Volume, 13=Value
    """
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    # English sheet: 'Bullient '  |  Arabic sheet: 'نشرة التداول'
    sheet_name = next((s for s in wb.sheetnames
                       if 'bull' in s.lower() or 'نشرة التداول' in s), None)
    if not sheet_name:
        print(f'    WARNING: unknown sheets {wb.sheetnames} in {date_iso}')
        return {}
    ws = wb[sheet_name]
    day_data = {}
    header_passed = False
    for row in ws.iter_rows(values_only=True):
        if not header_passed:
            # Header row: English 'Code' or Arabic 'رمز' in col 2
            if row[2] is not None and ('code' in str(row[2]).lower() or 'رمز' in str(row[2])):
                header_passed = True
            continue
        # Data rows: col 2 = ticker code (must be 2-6 uppercase letters)
        code = row[2]
        if not code or not isinstance(code, str):
            continue
        code = str(code).strip().upper()
        if not re.match(r'^[A-Z]{2,6}$', code):
            continue  # sector total rows or blank
        def f(v):
            try: return round(float(v), 4) if v is not None else None
            except: return None
        def i(v):
            try: return int(float(v)) if v is not None else 0
            except: return 0
        o, h, l, c = f(row[3]), f(row[4]), f(row[5]), f(row[8])
        deals, vol = i(row[11]), i(row[12])
        if c is None or c <= 0:
            continue
        day_data[code] = {'o': o, 'h': h, 'l': l, 'c': c, 'v': vol, 'd': deals}
    wb.close()
    return day_data

# ── Step 3: Download with retry ───────────────────────────────────────────────

def download_xlsx(url, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=H, timeout=30)
            if r.status_code == 200 and len(r.content) > 1000:
                return r.content
        except Exception as e:
            print(f'    Attempt {attempt+1} failed: {e}')
        time.sleep(1)
    return None

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    os.makedirs('data', exist_ok=True)

    # Load existing ohlcv.json if present
    ohlcv_path = 'data/ohlcv.json'
    if os.path.exists(ohlcv_path):
        with open(ohlcv_path) as f:
            ohlcv = json.load(f)
        print(f'Loaded existing ohlcv.json: {len(ohlcv)} dates')
    else:
        ohlcv = {}

    print('Scraping ISX portal pages...')
    files = get_all_daily_xlsx()
    print(f'Found {len(files)} daily XLSX files')

    # Sort by date
    files.sort(key=lambda x: x[0])

    new_dates = 0
    for date_iso, url in files:
        if date_iso in ohlcv:
            print(f'  Skip {date_iso} (already have)')
            continue
        print(f'  Downloading {date_iso}: {url.split("/")[-1][:40]}...')
        content = download_xlsx(url)
        if not content:
            print(f'    FAILED to download {url}')
            continue
        day_data = parse_xlsx(content, date_iso)
        if not day_data:
            print(f'    No data parsed for {date_iso}')
            continue
        ohlcv[date_iso] = day_data
        new_dates += 1
        print(f'    {len(day_data)} companies')
        time.sleep(0.4)   # be polite

    # Sort by date and save
    ohlcv_sorted = dict(sorted(ohlcv.items()))
    with open(ohlcv_path, 'w') as f:
        json.dump(ohlcv_sorted, f, separators=(',', ':'))
    print(f'\nohlcv.json written — {len(ohlcv_sorted)} dates, {new_dates} new')

    # ── Also backfill hist.json with closing prices ──────────────────────────
    # hist.json 's' dict: {SYM: [[ts, close], ...]}
    hist_path = 'data/hist.json'
    if os.path.exists(hist_path):
        with open(hist_path) as f:
            hist = json.load(f)
    else:
        hist = {'s': {}, 'l': {}}

    baghdad = datetime.timezone(datetime.timedelta(hours=3))
    added = 0
    for date_iso, day_data in ohlcv_sorted.items():
        y, m, d = map(int, date_iso.split('-'))
        dt = datetime.datetime(y, m, d, tzinfo=baghdad)
        if dt.weekday() in (4, 5):   # Fri=4, Sat=5 — skip weekends
            continue
        ts = int(dt.astimezone(datetime.timezone.utc).timestamp())
        for sym, row in day_data.items():
            arr = hist['s'].get(sym, [])
            if not any(p[0] == ts for p in arr):
                arr.append([ts, row['c']])
                arr.sort(key=lambda x: x[0])
                # deduplicate
                arr = [v for i, v in enumerate(arr) if i == 0 or v[0] != arr[i-1][0]]
                hist['s'][sym] = arr
                added += 1
            arr_l = hist.get('l', {}).get(sym, [])
            if not any(p[0] == ts for p in arr_l):
                arr_l.append([ts, row['c']])
                arr_l.sort(key=lambda x: x[0])
                arr_l = [v for i, v in enumerate(arr_l) if i == 0 or v[0] != arr_l[i-1][0]]
                hist.setdefault('l', {})[sym] = arr_l

    with open(hist_path, 'w') as f:
        json.dump(hist, f, separators=(',', ':'))
    print(f'hist.json updated — {added} new closing-price points added')

if __name__ == '__main__':
    main()
