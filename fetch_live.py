import requests, json, datetime, os
from concurrent.futures import ThreadPoolExecutor, as_completed

H = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Origin': 'https://rs.iq',
    'Referer': 'https://rs.iq/',
    'rb-lang': '1',
    'Accept': 'application/json, text/plain, */*',
}

stocks = requests.get('https://appapi.rs.iq/api/SiteStock/StocksList', headers=H, timeout=20).json()
print(f"StocksList: {len(stocks)} stocks")

def fetch_detail(s):
    try:
        d = requests.get(
            f'https://appapi.rs.iq/api/SiteStock/StockdetailsById?StockId={s["StockID"]}',
            headers=H, timeout=20
        ).json()
        cp = d.get('ClosingPrice') or 0
        if cp <= 0:
            return None
        op = d.get('OpeningPrice') or d.get('OpenPrice') or cp
        hi = d.get('HighPrice') or d.get('HighestPrice') or d.get('DayHigh') or cp
        lo = d.get('LowPrice') or d.get('LowestPrice') or d.get('DayLow') or cp
        deals = int(d.get('TradingDeals') or d.get('NumberOfDeals') or d.get('Deals') or 0)
        return {
            'code': s['StockCode'],
            'close': cp,
            'open':  round(op, 4),
            'high':  round(hi, 4),
            'low':   round(lo, 4),
            'change': round(d.get('DDChange') or 0, 4),
            'pct': round(s.get('DTDPriceChange') or 0, 4),
            'vol': int(d.get('TradingVolume') or 0),
            'deals': deals,
        }
    except Exception as e:
        print(f"  Skip {s['StockCode']}: {e}")
        return None

live_stocks = []
with ThreadPoolExecutor(max_workers=12) as ex:
    for r in as_completed([ex.submit(fetch_detail, s) for s in stocks]):
        v = r.result()
        if v:
            live_stocks.append(v)

print(f"Prices fetched: {len(live_stocks)}")

rsisx = requests.get('https://appapi.rs.iq/api/SiteStock/GetRSISXList?type=RSISX', headers=H, timeout=20).json()
latest = rsisx[-1] if rsisx else None
prev = rsisx[-2] if len(rsisx) >= 2 else None

rsisx_val = float(latest['IQD']) if latest else 0
rsisx_prev = float(prev['IQD']) if prev else 0
rsisx_change = round(rsisx_val - rsisx_prev, 2) if rsisx_prev else 0
rsisx_pct = round((rsisx_change / rsisx_prev) * 100, 2) if rsisx_prev else 0

# Market breadth from DTDPriceChange
up = sum(1 for s in stocks if (s.get('DTDPriceChange') or 0) > 0)
dn = sum(1 for s in stocks if (s.get('DTDPriceChange') or 0) < 0)
fl = len(stocks) - up - dn

# Sector % changes — average DTDPriceChange per sector
SECTOR_MAP = {
    'Banking Sector': 'BANK',
    'Insurance Sector': 'INS',
    'Investment Sector': 'INV',
    'Services Sector': 'SVC',
    'Industry Sector': 'IND',
    'Hotels&Tourism Sector': 'HTL',
    'Agriculture Sector': 'AGR',
    'Telecom Sector': 'TEL',
}
sector_data = {}
for s in stocks:
    sid = SECTOR_MAP.get(s.get('Sector', ''))
    if not sid: continue
    pct = s.get('DTDPriceChange') or 0
    if sid not in sector_data: sector_data[sid] = []
    sector_data[sid].append(pct)
sector_chg = {k: round(sum(v)/len(v), 2) for k, v in sector_data.items() if v}

os.makedirs('data', exist_ok=True)
out = {
    'updated': datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
    'stocks': live_stocks,
    'rsisx': {'value': latest['IQD'], 'date': latest['Date'], 'change': rsisx_change, 'pct': rsisx_pct} if latest else None,
    'breadth': {'up': up, 'dn': dn, 'fl': fl},
    'sectors': sector_chg
}
with open('public/data/live.json', 'w') as f:
    json.dump(out, f, separators=(',', ':'))
print(f"live.json written — {len(live_stocks)} stocks, RSISX={out['rsisx']}")

# ── Update hist.json with today's closing prices ──
# Timestamp convention: midnight Baghdad (UTC+3) = 21:00 UTC previous day
baghdad = datetime.timezone(datetime.timedelta(hours=3))
now_baghdad = datetime.datetime.now(baghdad)
today_midnight_baghdad = now_baghdad.replace(hour=0, minute=0, second=0, microsecond=0)
today_ts = int(today_midnight_baghdad.astimezone(datetime.timezone.utc).timestamp())

# Skip on weekends (Fri=4, Sat=5 in Baghdad)
weekday = today_midnight_baghdad.weekday()  # Mon=0 … Sun=6
if weekday in (4, 5):  # Friday or Saturday
    print(f"Weekend ({today_midnight_baghdad.strftime('%A')}) — skipping hist update")
else:
    hist_path = 'public/data/hist.json'
    if os.path.exists(hist_path):
        with open(hist_path) as f:
            hist = json.load(f)
    else:
        hist = {'s': {}, 'l': {}}

    price_map = {s['code']: s['close'] for s in live_stocks}
    updated_count = 0

    for sym, price in price_map.items():
        if price <= 0:
            continue
        for key, max_pts in [('s', 252), ('l', None)]:
            arr = hist[key].get(sym, [])
            # Only append if this timestamp isn't already there
            if not any(p[0] == today_ts for p in arr):
                arr.append([today_ts, round(price, 4)])
                arr.sort(key=lambda x: x[0])
                # deduplicate
                arr = [v for i, v in enumerate(arr) if i == 0 or v[0] != arr[i-1][0]]
                if max_pts and len(arr) > max_pts:
                    arr = arr[-max_pts:]
                hist[key][sym] = arr
                if key == 's':
                    updated_count += 1

    # ── Update RSISX history in hist.json ──
    # Store full RSISX series from API (always replace to stay in sync)
    def parse_rsisx_date(d):
        # d.Date is "M/D/YYYY"
        parts = d['Date'].split('/')
        m, day, y = int(parts[0]), int(parts[1]), int(parts[2])
        dt = datetime.datetime(y, m, day, tzinfo=baghdad)
        return int(dt.astimezone(datetime.timezone.utc).timestamp())

    rsisx_pts = []
    for r in rsisx:
        try:
            ts = parse_rsisx_date(r)
            rsisx_pts.append([ts, float(r['IQD'])])
        except Exception:
            continue

    hist['rsisx_s'] = rsisx_pts[-365:] if len(rsisx_pts) > 365 else rsisx_pts
    hist['rsisx_l'] = rsisx_pts

    with open(hist_path, 'w') as f:
        json.dump(hist, f, separators=(',', ':'))
    print(f"hist.json updated — {updated_count} symbols, {len(rsisx_pts)} RSISX pts for {today_midnight_baghdad.date()}")

# ── Append today's OHLCV from ISX portal daily XLSX ──────────────────────────
import re, time
from io import BytesIO
import openpyxl

today_iso = today_midnight_baghdad.strftime('%Y-%m-%d')
ohlcv_path = 'public/data/ohlcv.json'
ohlcv = json.load(open(ohlcv_path)) if os.path.exists(ohlcv_path) else {}

day_data = {}  # populated below — used to sync live.json after XLSX is available

if today_iso in ohlcv:
    day_data = ohlcv[today_iso]   # already fetched in a previous run — still sync live.json
    print(f"ohlcv: {today_iso} already present")
elif weekday not in (4, 5):
    DOMAIN = 'http://www.isx-iq.net'
    PH = {'User-Agent': 'Mozilla/5.0'}
    try:
        r = requests.get(f'{DOMAIN}/isxportal/portal/uploadedFilesList.html', headers=PH, timeout=20)
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', r.text, re.DOTALL)
        xlsx_url = None
        for row in rows:
            if '.xlsx' not in row: continue
            lm = re.search(r'href="(/isxportal/files/([^"]+\.xlsx))"', row)
            dm = re.search(r'\b(\d{2}/\d{2}/\d{4})\b', row)
            if not lm or not dm: continue
            fname = lm.group(2)
            if not fname or ord(fname[0]) > 0x05FF: continue  # skip Arabic-named files
            d, m, y = dm.group(1).split('/')
            if f'{y}-{m}-{d}' == today_iso:
                xlsx_url = DOMAIN + lm.group(1)
                break
        if xlsx_url:
            print(f"ohlcv: downloading today's XLSX: {xlsx_url.split('/')[-1][:40]}")
            rc = requests.get(xlsx_url, headers=PH, timeout=30)
            if rc.status_code == 200 and len(rc.content) > 1000:
                wb = openpyxl.load_workbook(BytesIO(rc.content), read_only=True, data_only=True)
                sn = next((s for s in wb.sheetnames if 'bull' in s.lower() or 'نشرة التداول' in s), None)
                if sn:
                    day_data = {}
                    header_passed = False
                    for row in wb[sn].iter_rows(values_only=True):
                        if not header_passed:
                            if row[2] is not None and ('code' in str(row[2]).lower() or 'رمز' in str(row[2])):
                                header_passed = True
                            continue
                        code = row[2]
                        if not code or not isinstance(code, str): continue
                        code = str(code).strip().upper()
                        if not re.match(r'^[A-Z]{2,6}$', code): continue
                        def _f(v):
                            try: return round(float(v), 4) if v is not None else None
                            except: return None
                        c = _f(row[8])
                        if not c or c <= 0: continue
                        day_data[code] = {'o':_f(row[3]),'h':_f(row[4]),'l':_f(row[5]),'c':c,
                                          'v':int(float(row[12] or 0)),'d':int(float(row[11] or 0))}
                    if day_data:
                        ohlcv[today_iso] = day_data
                        with open(ohlcv_path, 'w') as f:
                            json.dump(dict(sorted(ohlcv.items())), f, separators=(',', ':'))
                        print(f"ohlcv.json updated — {len(day_data)} companies for {today_iso}")
                    else:
                        print("ohlcv: no data parsed from today's XLSX")
        else:
            print(f"ohlcv: today's XLSX ({today_iso}) not yet posted on ISX portal")
    except Exception as e:
        print(f"ohlcv: error fetching today's XLSX: {e}")

# ── Sync live.json prices with ISX XLSX (authoritative source) ───────────────
# The Rabee API often returns yesterday's ClosingPrice during/after trading.
# The ISX daily XLSX is the official settlement price — always prefer it.
if day_data and weekday not in (4, 5):
    # Build previous-close map from hist.json so we can recalculate pct change
    prev_close_map = {}
    try:
        hd = json.load(open('public/data/hist.json'))
        for sym, pts in hd.get('s', {}).items():
            # today_ts = midnight Baghdad — filter to strictly before today
            prev = [p for p in pts if p[0] < today_ts]
            if prev:
                prev_close_map[sym] = prev[-1][1]
    except Exception as e:
        print(f"prev-close lookup failed: {e}")

    override_count = 0
    for stock in out['stocks']:
        code = stock['code']
        if code not in day_data:
            continue
        xd = day_data[code]
        nc = xd.get('c')
        if not nc or nc <= 0:
            continue
        prev = prev_close_map.get(code)
        chg  = round(nc - prev, 4) if prev else stock['change']
        pct  = round((chg / prev) * 100, 4) if prev else stock['pct']
        stock.update({
            'close':  nc,
            'open':   xd.get('o') or stock['open'],
            'high':   xd.get('h') or stock['high'],
            'low':    xd.get('l') or stock['low'],
            'vol':    xd.get('v') or stock['vol'],
            'deals':  xd.get('d') or stock['deals'],
            'change': chg,
            'pct':    pct,
        })
        override_count += 1

    if override_count:
        with open('public/data/live.json', 'w') as f:
            json.dump(out, f, separators=(',', ':'))
        print(f"live.json synced with XLSX: {override_count} prices overridden (ISX official > Rabee API)")
