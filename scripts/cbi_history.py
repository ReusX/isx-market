#!/usr/bin/env python3
"""
Parse the Central Bank's daily USD/IQD workbook into {date: rate} JSON.

    python3 scripts/cbi_history.py <listing-url>   # emits JSON on stdout

Why Python here, in a TypeScript repo: this file is read by openpyxl, which
is already a dependency of the ingestion layer and is the only reader whose
output has been verified cell-by-cell against the workbook.

A hand-rolled XML reader was tried first and produced 2,083 missing days and
1,996 days that are not in the workbook at all — with every value it *did*
emit matching exactly, which is what made the fault so hard to see. Values
were right; dates were wrong. Parsing a real spreadsheet format with regular
expressions was the mistake, not any single bug in that code.

── The sheet geometry ─────────────────────────────────────────────────────
Not a table of dates. Each year is a block: a header cell reading `2003/Date`
in column A, month names across columns B..M, and **day N exactly N rows below
the header**. Column A of each data row repeats the day, which is used to
confirm alignment rather than trusting the offset. Blank cells are non-trading
days and are skipped — never carried forward, because an interpolated official
rate is a value the bank never published.

The workbook URL is not stable: files live at /static/uploads/up/file-<id>.xlsx
and the id changes on every re-upload, so the listing page is scraped for the
current link each run.
"""
import datetime
import io
import json
import re
import sys
import urllib.request

import openpyxl

UA = {"User-Agent": "Mozilla/5.0 (compatible; IraqSM/1.0; +https://iraqsm.com)"}
MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}
SHEET = re.compile(r"daily price basis", re.I)
HEADER = re.compile(r"^((?:19|20)\d{2})/Date")


def workbook_urls(listing: str) -> list[str]:
    req = urllib.request.Request(listing, headers=UA)
    html = urllib.request.urlopen(req, timeout=30).read().decode("utf-8", "ignore")
    urls = re.findall(r'href="(https://cbi\.iq/static/uploads/up/[^"]+\.xlsx)"', html)
    if not urls:
        raise SystemExit("no .xlsx links on the listing page — its structure changed")
    return list(dict.fromkeys(urls))


def parse(buf: bytes) -> dict[str, float]:
    wb = openpyxl.load_workbook(io.BytesIO(buf), data_only=True, read_only=True)
    names = [n for n in wb.sheetnames if SHEET.search(n)]
    if not names:
        return {}
    ws = wb[names[0]]
    rows = [list(r) for r in ws.iter_rows(max_col=13, values_only=True)]

    out: dict[str, float] = {}
    for i, row in enumerate(rows):
        head = HEADER.match(str(row[0])) if row[0] is not None else None
        if not head:
            continue
        year = int(head.group(1))
        months = {
            c: MONTHS[str(row[c] or "").strip().lower()[:3]]
            for c in range(1, len(row))
            if str(row[c] or "").strip().lower()[:3] in MONTHS
        }
        if not months:
            continue
        for day in range(1, 32):
            if i + day >= len(rows):
                break
            data = rows[i + day]
            # Column A must confirm the day, so a shifted or truncated block is
            # skipped rather than silently misread against the wrong year.
            if data[0] != day:
                continue
            for col, month in months.items():
                v = data[col] if col < len(data) else None
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                if not (100 < float(v) < 5000):
                    continue
                try:
                    datetime.date(year, month, day)
                except ValueError:
                    continue          # 31 February is a blank, not an observation
                out[f"{year:04d}-{month:02d}-{day:02d}"] = float(v)
    return out


def main() -> None:
    listing = sys.argv[1] if len(sys.argv) > 1 else "https://cbi.iq/page/144"
    merged: dict[str, float] = {}
    for url in workbook_urls(listing):
        try:
            req = urllib.request.Request(url, headers=UA)
            buf = urllib.request.urlopen(req, timeout=120).read()
            found = parse(buf)
        except Exception as exc:                      # noqa: BLE001
            print(f"  · {url[-28:]} → {exc}", file=sys.stderr)
            continue
        print(f"  {'✓' if found else '·'} {url[-28:]} → {len(found)} days", file=sys.stderr)
        merged.update(found)
    if not merged:
        raise SystemExit("no observations extracted")
    json.dump(merged, sys.stdout)


if __name__ == "__main__":
    main()
